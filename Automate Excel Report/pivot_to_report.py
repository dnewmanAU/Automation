import os, sys
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

app_path = os.path.dirname(sys.executable)

input_path = os.path.join(app_path, "pivot-table.xlsx")
wb = load_workbook(input_path)
sheet = wb["Report"]
bar_chart = BarChart()
month_title = input("Which month: ")

# Get the indices of columns and rows
min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Select cells on the left excluding the header
categories = Reference(
    sheet, min_col=min_col, max_col=min_col, min_row=min_row + 1, max_row=max_row
)
# Select everything else
data = Reference(
    sheet, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row
)

# Add the bar chart
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
bar_chart.title = "Sales by Product line"
bar_chart.style = 13
sheet.add_chart(bar_chart, "B12")  # put it at cell B-12

# Loop over and put the sum of the columns in a new row
for i in range(min_col + 1, max_col + 1):
    letter = get_column_letter(i)  # convert index to letter
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"

# Title the report
sheet["A1"] = "Sales Report"
sheet["A2"] = month_title
sheet["A1"].font = Font("Arial", bold=True, size=20)
sheet["A2"].font = Font("Arial", bold=True, size=10)

output_path = os.path.join(app_path, f"report-{month_title}.xlsx")

wb.save(output_path)
